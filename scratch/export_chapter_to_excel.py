import os
import sys
import json
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Load .env
dotenv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env')
load_dotenv(dotenv_path)

db_url = os.getenv("DATABASE_URL")
engine = create_engine(db_url)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

CLR_HEADER_BG  = "1A365D"  # Navy blue
CLR_HEADER_FG  = "FFFFFF"
CLR_CHAPTER_BG = "2B6CB0"  # Medium blue
CLR_CHAPTER_FG = "FFFFFF"
CLR_ALT_ROW    = "EDF2F7"  # Very light greyish-blue
CLR_WHITE      = "FFFFFF"
CLR_TITLE_BG   = "2A4365"

THIN_SIDE   = Side(style="thin")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

COLUMNS = [
    ("#",                   5),
    ("Purpose",            14),
    ("Topic",              30),
    ("Learning Unit",      30),
    ("Question Type",      14),
    ("Question",           65),
    ("MCQ Options",        40),
    ("Correct Answer",     22),
    ("Acceptable Answers", 35),
    ("Hint 1",             40),
    ("Hint 2",             40),
    ("Explanation",        50),
]

def acceptable_str(raw):
    if raw is None:
        return "—"
    if isinstance(raw, list):
        return ", ".join(str(a) for a in raw) if raw else "—"
    if isinstance(raw, str):
        try:
            parsed = json.loads(raw)
            return ", ".join(str(a) for a in parsed) if parsed else "—"
        except Exception:
            return raw
    return str(raw)

def mcq_options_str(raw):
    if raw is None:
        return ""
    opts = raw if isinstance(raw, list) else json.loads(raw) if isinstance(raw, str) else []
    if not opts:
        return ""
    labels = "ABCDE"
    return "  ".join(f"{labels[i]}) {o}" for i, o in enumerate(opts))

def export_chapter_to_excel(ch_title, output_path):
    print(f"Fetching questions for chapter '{ch_title}'...")
    
    query = text("""
        SELECT
            bo.name                         AS board,
            gr.name                         AS grade,
            su.name                         AS subject,
            ch.title                        AS chapter,
            tp.title                        AS topic,
            lu.title                        AS learning_unit,
            q.question_type,
            q.text                          AS question_text,
            q.mcq_options,
            q.correct_option,
            q.acceptable_answers,
            q.hint_level_1,
            q.hint_level_2,
            q.full_explanation,
            q.question_purpose
        FROM questions q
        JOIN learning_units lu  ON lu.id  = q.learning_unit_id
        JOIN subtopics      st  ON st.id  = lu.subtopic_id
        JOIN topics         tp  ON tp.id  = st.topic_id
        JOIN chapters       ch  ON ch.id  = tp.chapter_id
        JOIN subjects       su  ON su.id  = ch.subject_id
        JOIN grades         gr  ON gr.id  = su.grade_id
        JOIN boards         bo  ON bo.id  = gr.board_id
        WHERE ch.title = :title
        ORDER BY tp.title, lu.title, q.created_at;
    """)
    
    with engine.connect() as conn:
        res = conn.execute(query, {"title": ch_title})
        rows = res.fetchall()
        
    if not rows:
        print(f"No questions found for chapter '{ch_title}'. Skipping export.")
        return
        
    print(f"Found {len(rows)} questions. Generating Excel file at {output_path}...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Question Bank"
    ws.views.sheetView[0].showGridLines = True
    
    # Write Main Title
    ws.merge_cells("A1:L1")
    ws["A1"] = f"QUESTION BANK: {ch_title.upper()}"
    ws["A1"].font = Font(name="Segoe UI", size=16, bold=True, color=CLR_HEADER_FG)
    ws["A1"].fill = PatternFill(start_color=CLR_TITLE_BG, end_color=CLR_TITLE_BG, fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Write Headers
    for col_idx, (col_name, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color=CLR_HEADER_FG)
        cell.fill = PatternFill(start_color=CLR_HEADER_BG, end_color=CLR_HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 28
    
    # Write Rows
    for row_idx, r in enumerate(rows, 3):
        ws.row_dimensions[row_idx].height = 24
        fill_color = CLR_ALT_ROW if row_idx % 2 == 0 else CLR_WHITE
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # #
        c = ws.cell(row=row_idx, column=1, value=row_idx - 2)
        c.font = Font(name="Segoe UI", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        
        # Purpose
        c = ws.cell(row=row_idx, column=2, value=str(r[14]) if r[14] else "—")
        c.font = Font(name="Segoe UI", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        
        # Topic
        c = ws.cell(row=row_idx, column=3, value=r[4])
        c.font = Font(name="Segoe UI", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Learning Unit
        c = ws.cell(row=row_idx, column=4, value=r[5])
        c.font = Font(name="Segoe UI", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Question Type
        c = ws.cell(row=row_idx, column=5, value=r[6])
        c.font = Font(name="Segoe UI", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        
        # Question Text
        c = ws.cell(row=row_idx, column=6, value=r[7])
        c.font = Font(name="Segoe UI", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # MCQ Options
        c = ws.cell(row=row_idx, column=7, value=mcq_options_str(r[8]))
        c.font = Font(name="Segoe UI", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Correct Answer
        c = ws.cell(row=row_idx, column=8, value=str(r[9]) if r[9] is not None else "—")
        c.font = Font(name="Segoe UI", size=10, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        
        # Acceptable Answers
        c = ws.cell(row=row_idx, column=9, value=acceptable_str(r[10]))
        c.font = Font(name="Segoe UI", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Hint 1
        c = ws.cell(row=row_idx, column=10, value=r[11] if r[11] else "—")
        c.font = Font(name="Segoe UI", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Hint 2
        c = ws.cell(row=row_idx, column=11, value=r[12] if r[12] else "—")
        c.font = Font(name="Segoe UI", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Explanation
        c = ws.cell(row=row_idx, column=12, value=r[13] if r[13] else "—")
        c.font = Font(name="Segoe UI", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        for col_idx in range(1, 13):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            
    # Set Column Widths
    for col_idx, (_, width) in enumerate(COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        
    wb.save(output_path)
    print(f"Successfully saved Question Bank Excel to: {output_path}")

if __name__ == "__main__":
    ch_map = {
        "How I Taught My Grandmother to Read": "Chapter_2_Question_Bank_Review.xlsx",
        "The Incident of the Tooth": "Chapter_3_Question_Bank_Review.xlsx",
        "Children of India": "Chapter_4_Question_Bank_Review.xlsx",
        "सोनकंठी गौरैया": "Chapter_3_Hindi_Question_Bank_Review.xlsx",
        "ननिहाल": "Chapter_4_Hindi_Question_Bank_Review.xlsx",
        "संज्ञा": "Chapter_8_Hindi_Question_Bank_Review.xlsx",
        "संज्ञा के विकारिक तत्व": "Chapter_9_Hindi_Question_Bank_Review.xlsx",
        "Exploring Magnets": "Chapter_4_Science_Question_Bank_Review.xlsx",
        "Measurement of Length and Motion": "Chapter_5_Science_Question_Bank_Review.xlsx",
        "Timeline and Sources of History India": "Chapter_4_Social_Science_Question_Bank_Review.xlsx"
    }
    for ch_title, filename in ch_map.items():
        export_chapter_to_excel(
            ch_title,
            f"generated/questions/{filename}"
        )
