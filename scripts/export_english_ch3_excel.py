import os
import sys
import json
from datetime import datetime

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from sqlalchemy import create_engine, text
from app.core.config import settings

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Styles for English Theme (Blue/Navy Theme)
CLR_HEADER_BG  = "1A365D"  # Navy blue
CLR_HEADER_FG  = "FFFFFF"
CLR_CHAPTER_BG = "2B6CB0"  # Medium blue
CLR_CHAPTER_FG = "FFFFFF"
CLR_ALT_ROW    = "EDF2F7"  # Very light greyish-blue
CLR_WHITE      = "FFFFFF"
CLR_TITLE_BG   = "2A4365"

THIN_SIDE   = Side(style="thin")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

COLUMNS = [
    ("#",                   5),
    ("Purpose",            14),
    ("Topic",              30),
    ("Learning Unit",      30),
    ("Question Type",      14),
    ("Question",           65),
    ("MCQ Options",        40),
    ("Correct Answer",     22),
    ("Acceptable Answers", 35),
    ("Hint 1",             40),
]

def acceptable_str(raw):
    if raw is None:
        return "—"
    if isinstance(raw, list):
        return ", ".join(str(a) for a in raw) if raw else "—"
    if isinstance(raw, str):
        try:
            parsed = json.loads(raw)
            return ", ".join(str(a) for a in parsed) if parsed else "—"
        except Exception:
            return raw
    return str(raw)

def mcq_options_str(raw):
    if raw is None:
        return ""
    opts = raw if isinstance(raw, list) else json.loads(raw) if isinstance(raw, str) else []
    if not opts:
        return ""
    labels = "ABCDE"
    return "  ".join(f"{labels[i]}) {o}" for i, o in enumerate(opts))

def export_chapter_to_excel(chapter_id, output_path):
    engine = create_engine(settings.DATABASE_URL)
    
    query = text("""
        SELECT
            bo.name                         AS board,
            gr.name                         AS grade,
            su.name                         AS subject,
            ch.title                        AS chapter,
            tp.title                        AS topic,
            lu.title                        AS learning_unit,
            q.question_type,
            q.text                          AS question_text,
            q.mcq_options,
            q.correct_option,
            q.acceptable_answers,
            q.hint_level_1,
            q.hint_level_2,
            q.question_purpose
        FROM questions q
        JOIN learning_units lu  ON lu.id  = q.learning_unit_id
        JOIN subtopics      st  ON st.id  = lu.subtopic_id
        JOIN topics         tp  ON tp.id  = st.topic_id
        JOIN chapters       ch  ON ch.id  = tp.chapter_id
        JOIN subjects       su  ON su.id  = ch.subject_id
        JOIN grades         gr  ON gr.id  = su.grade_id
        JOIN boards         bo  ON bo.id  = gr.board_id
        WHERE tp.chapter_id = :chapter_id
        ORDER BY tp.title, lu.title, q.created_at;
    """)
    
    with engine.connect() as conn:
        rows = conn.execute(query, {"chapter_id": chapter_id}).fetchall()
        
    if not rows:
        print(f"No questions found for chapter ID {chapter_id}")
        return False
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review Questions"
    
    # 1. Title Block
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value=f"{rows[0].board} | Grade {rows[0].grade} | {rows[0].subject}")
    title_cell.font = Font(bold=True, size=14, color=CLR_HEADER_FG)
    title_cell.fill = PatternFill("solid", fgColor=CLR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # Chapter Block
    ws.merge_cells("A2:H2")
    ch_cell = ws.cell(row=2, column=1, value=f"Chapter: {rows[0].chapter}")
    ch_cell.font = Font(bold=True, size=12, color=CLR_CHAPTER_FG)
    ch_cell.fill = PatternFill("solid", fgColor=CLR_CHAPTER_BG)
    ch_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28
    
    # Headers
    row_idx = 4
    ws.row_dimensions[row_idx].height = 28
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color=CLR_HEADER_FG, size=11)
        cell.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
        
    # Data rows
    serial = 1
    alt_row = False
    for r in rows:
        row_idx += 1
        ws.row_dimensions[row_idx].height = 45 # default height
        fill = PatternFill("solid", fgColor=CLR_ALT_ROW if alt_row else CLR_WHITE)
        
        is_mcq = str(r.question_type).upper() in ("MCQ", "RECALL") and r.mcq_options
        
        values = [
            serial,
            r.question_purpose or "—",
            r.topic or "—",
            r.learning_unit or "—",
            r.question_type,
            r.question_text,
            mcq_options_str(r.mcq_options) if is_mcq else "",
            r.correct_option or "" if is_mcq else r.expected_answer if hasattr(r, 'expected_answer') else r.correct_option or "",
            acceptable_str(r.acceptable_answers),
            r.hint_level_1 or "—",
        ]
        
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 8 and is_mcq and val:
                cell.font = Font(bold=True, color="1B5E20")  # Green for answers
            elif col_idx == 9:
                cell.font = Font(color="2D3748")
                
        serial += 1
        alt_row = not alt_row
        
    wb.save(output_path)
    print(f"Successfully exported {len(rows)} questions to {output_path}")
    return True

if __name__ == "__main__":
    chapter_id = sys.argv[1] if len(sys.argv) > 1 else "47421526-d2a9-46e4-a528-a17f038ba5d3"
    if chapter_id == "47421526-d2a9-46e4-a528-a17f038ba5d3":
        filename = "english_chapter_3_review_questions.xlsx"
    elif chapter_id == "77cb4faf-00a6-42e3-9eb6-ba4d561b510f":
        filename = "english_chapter_4_review_questions.xlsx"
    elif chapter_id == "ac98c7ca-eb95-40b7-bb3d-f1d5742b129c":
        filename = "hindi_chapter_4_review_questions.xlsx"
    elif chapter_id == "afacd67d-cf44-4d9a-850c-de0e11c3fb1a":
        filename = "hindi_chapter_8_review_questions.xlsx"
    elif chapter_id == "18ccdb00-bad8-4e7e-af14-532d558cccf7":
        filename = "hindi_chapter_9_review_questions.xlsx"
    elif chapter_id == "6aef1a7f-cb2e-40be-8f16-2f9c770ec866":
        filename = "science_chapter_5_review_questions.xlsx"
    elif chapter_id == "b36776b5-a9bf-4260-afac-261285992222":
        filename = "social_science_chapter_4_review_questions.xlsx"
    elif chapter_id in ["3027d105-1c4e-49d8-a76c-50979075e841", "e76b1d44-5008-4170-b283-607e77a37adb"]:
        filename = "social_science_chapter_10_review_questions.xlsx"
    elif chapter_id == "bf2b946f-f190-4ede-a26c-e657403d5132":
        filename = "gujarati_chapter_3_review_questions.xlsx"
    else:
        filename = "english_chapter_2_review_questions.xlsx"
    export_chapter_to_excel(chapter_id, filename)
